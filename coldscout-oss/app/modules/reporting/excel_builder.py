"""Excel report builder for Cold Scout OSS."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
from typing import List, Dict, Any


def generate_daily_report_excel(report_data: Dict[str, Any], leads: List[Dict[str, Any]], output_date: date) -> str:
    wb = openpyxl.Workbook()
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    data_font = Font(name="Calibri", size=11, color="000000")
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    border_bottom = Border(bottom=Side(style="thin", color="EAEAEA"))

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Daily Summary"
    ws1.merge_cells("A1:B1")
    ws1.cell(row=1, column=1, value=f"Cold Scout  ·  Report  ·  {output_date}").font = Font(bold=True, size=14, color="FFFFFF")
    ws1.cell(row=1, column=1).fill = header_fill

    for i, h in enumerate(["Metric", "Value"]):
        c = ws1.cell(row=3, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill

    metrics = [
        ("Leads discovered", report_data.get('leads_discovered', 0)),
        ("Qualified", report_data.get('leads_qualified', 0)),
        ("Emails sent", report_data.get('emails_sent', 0)),
        ("Opened", report_data.get('emails_opened', 0)),
        ("Clicked", report_data.get('links_clicked', 0)),
        ("Replies", report_data.get('replies_received', 0)),
    ]
    for i, (label, value) in enumerate(metrics):
        row = 4 + i
        ws1.cell(row=row, column=1, value=label).font = data_font
        ws1.cell(row=row, column=2, value=value).font = Font(bold=True, size=12)
        ws1.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            ws1.cell(row=row, column=1).fill = alt_fill
            ws1.cell(row=row, column=2).fill = alt_fill

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 16

    # Sheet 2: Lead Details
    ws2 = wb.create_sheet("Lead Details")
    headers = ["Business", "Category", "City", "Status", "Email Sent", "Phone", "Google Maps"]
    for i, h in enumerate(headers):
        c = ws2.cell(row=1, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill

    for ri, lead in enumerate(leads):
        row = 2 + ri
        values = [
            lead.get("business_name"), lead.get("category"), lead.get("city"),
            lead.get("status"), "Yes" if lead.get("email_sent_at") else "No",
            lead.get("phone"), lead.get("google_maps_url"),
        ]
        for ci, val in enumerate(values):
            c = ws2.cell(row=row, column=ci + 1, value=val)
            c.font = data_font
            if ri % 2 == 0:
                c.fill = alt_fill

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20

    os.makedirs("tmp", exist_ok=True)
    filepath = os.path.join("tmp", f"ColdScout_Report_{output_date}.xlsx")
    wb.save(filepath)
    return filepath
