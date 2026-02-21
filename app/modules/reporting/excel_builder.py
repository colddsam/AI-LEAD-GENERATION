"""
Excel report generation module.
Utilizes the openpyxl library to systematically assemble quantitative metrics
and lead status details into a formatted administrative spreadsheet artifact.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import date
import os
from typing import List, Dict, Any

def generate_daily_report_excel(report_data: Dict[str, Any], leads: List[Dict[str, Any]], output_date: date) -> str:
    """
    Compiles daily performance metrics and individual lead properties into 
    a multi-sheet formatted Excel workbook.
    
    Args:
        report_data (Dict[str, Any]): Dictionary containing foundational daily metrics.
        leads (List[Dict[str, Any]]): Structurally formatted collection of recent lead data.
        output_date (date): The contextual date associated with the report.
        
    Returns:
        str: Expected file path to the synchronously generated Excel document.
    """
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Daily Summary"
    
    header_font = Font(bold=True, size=12)
    bg_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    ws1.append(["Metric", "Value"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = bg_fill
        
    ws1.append(["Total leads discovered", report_data.get('leads_discovered', 0)])
    ws1.append(["Qualified leads", report_data.get('leads_qualified', 0)])
    ws1.append(["Emails sent", report_data.get('emails_sent', 0)])
    ws1.append(["Emails opened", report_data.get('emails_opened', 0)])
    ws1.append(["Links clicked", report_data.get('links_clicked', 0)])
    ws1.append(["Replies received", report_data.get('replies_received', 0)])
    
    ws1.column_dimensions['A'].width = 25
    
    ws2 = wb.create_sheet("Lead Details")
    headers = ["Business", "Category", "Location", "Email Sent", "Opened", "Clicked", "Replied", "Status", "Phone", "Google Maps"]
    ws2.append(headers)
    
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = bg_fill
        
    for lead in leads:
        ws2.append([
            lead.get("business_name"),
            lead.get("category"),
            lead.get("city"),
            "Yes" if lead.get("email_sent_at") else "No",
            "Yes" if lead.get("first_opened_at") else "No",
            "Yes" if lead.get("first_clicked_at") else "No",
            "Yes" if lead.get("first_replied_at") else "No",
            lead.get("status"),
            lead.get("phone"),
            lead.get("google_maps_url")
        ])
        
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 20
        
    os.makedirs("tmp", exist_ok=True)
    filename = f"LeadGen_Report_{output_date.strftime('%Y-%m-%d')}.xlsx"
    filepath = os.path.join("tmp", filename)
    wb.save(filepath)
    
    return filepath
